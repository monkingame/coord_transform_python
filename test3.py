import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

excel_path=r'C:\Users\sun\Downloads\坐标-测试.xlsx'

df = pd.read_excel(excel_path,converters={'唯一编码':str})
# 加载 Excel 文件
book = load_workbook(excel_path)
# 获取第一个 sheet
ws = book.active

# 列名列表
columns = ['经营店铺名称', '唯一编码', '详细地址']

# 遍历 DataFrame，将数据写入到 Excel 文件中
for idx, row in df.iterrows():
    # 获取行索引，行索引从第三行开始（因为第一行是标题行，第二行是数据行）
    row_idx = idx + 2
    for col_name in columns:
        # print(col_name)
        # # 获取列索引
        col_idx = column_index_from_string(ws[col_name].column)
        print(col_idx)
        # # 获取单元格
        # cell = ws.cell(row=row_idx, column=col_idx)
        # # 将数据写入单元格
        # cell.value = row[col_name]

# 保存 Excel 文件
# book.save('data.xlsx')
